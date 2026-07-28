import re, os, sys
import openpyxl

if len(sys.argv) == 1:
    print("Usage: python term_extractor.py [INPUT WORKBOOK] [TARGET WORKBOOK]")
    exit()

# Getting the input

input_workbook = sys.argv[1]
target_workbook = sys.argv[2]

# Extracting the structure

workbook = openpyxl.load_workbook(target_workbook)

terms = {}

for sheet in workbook.worksheets:
    for row in sheet.rows:
        terms[row[0].value] = (row[1].value if len(row) > 1 else "", sheet.title)

# Extracting the values

workbook = openpyxl.load_workbook(input_workbook)

for sheet in workbook.worksheets:
    for row in sheet.rows:
        key = row[0].value
        value = row[1].value if len(row) > 1 else ""

        if not key in terms.keys(): continue

        terms[key] = (value, terms[key][1])

# Sorting the terms

terms = {term: value for term, value in sorted(terms.items(), key=lambda item: item[0])}


# Clearing the workbook

for sheet in workbook.worksheets:
    workbook.remove(sheet)

# Saving the data

for term, (value, sheet_name) in terms.items():
    if not sheet_name: sheet_name = "default"
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    sheet.append([term, value])


workbook.save(input_workbook)
