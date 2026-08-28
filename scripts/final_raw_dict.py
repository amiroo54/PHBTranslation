import re, os, sys
import openpyxl
from utils import pascal_to_natural

workbook_path = "dictionaries/FinalDict.xlsx"

steps = {
    "PHBFA/Chapter 7/Spell Descriptions": "Spells",
    "PHBFA/Chapter 5/Origin Feats/Origin Feats.md": "Feats",
    "PHBFA/Chapter 5/General Feats/General Feats.md": "Feats",
    "PHBFA/Chapter 5/Fighting Style Feats/Fighting Style Feats.md": "Feats",
    "PHBFA/Chapter 5/Epic Boon Feats/Epic Boon Feats.md": "Feats",
    "PHBFA/Chapter 4/Background Descriptions/Background Descriptions.md": "Origins",
    "PHBFA/Chapter 4/Species Descriptions/Species Descriptions.md": "Origins",
    "PHBFA/Chapter 1": "Basics",
    "PHBFA/Chapter 2": "Basics",
    "PHBFA/Chapter 6": "Equipment",
    "PHBFA/Chapter 4": "Origins",
    "PHBFA/Chapter 3/Barbarian": "Barbarian",
    "PHBFA/Chapter 3/Bard": "Bard",
    "PHBFA/Chapter 3/Cleric": "Cleric",
    "PHBFA/Chapter 3/Druid": "Druid",
    "PHBFA/Chapter 3/Fighter": "Fighter",
    "PHBFA/Chapter 3/Monk": "Monk",
    "PHBFA/Chapter 3/Paladin": "Paladin",
    "PHBFA/Chapter 3/Ranger": "Ranger",
    "PHBFA/Chapter 3/Rogue": "Rogue",
    "PHBFA/Chapter 3/Sorcerer": "Sorcerer",
    "PHBFA/Chapter 3/Warlock": "Warlock",
    "PHBFA/Chapter 3/Wizard": "Wizard",
    "PHBFA": "Basics"

}

for scan_path, sheet_name in steps.items():

    print(scan_path, sheet_name)

    terms = []
    directories = [scan_path]

    # Extracting the terms from the directory/file

    while directories:
        curr = directories.pop()
        if os.path.isdir(curr):
            for subdir in os.listdir(curr):
                directories.append(os.path.join(curr, subdir))
        if os.path.isfile(curr):
            if not curr.endswith(".md"):
                continue
            with open(curr, "r", encoding="utf-8") as file:
                matches = re.findall(r"\{\{(.*?)\}\}", file.read())
            for match in matches:
                parts = match.split(":")
                res = parts[0] + ":%" * (len(parts) - 1)
                terms.append(res)

    terms = dict.fromkeys(sorted(terms), ("", sheet_name))

    # Extracting the terms from the workbook

    workbook = openpyxl.Workbook()
    if workbook_path and os.path.exists(workbook_path):
        workbook = openpyxl.load_workbook(workbook_path)

    for sheet in workbook.worksheets:
        for row in sheet.rows:
            terms[row[0].value] = (row[1].value if len(row) > 1 else "", sheet.title)

    # Clearaing the sheets

    for sheet in workbook.worksheets:
        workbook.remove(sheet)

    # Adding the terms to the workbook

    for term, (value, sheet_name) in terms.items():
        if not sheet_name: sheet_name = "default"
        
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]

        sheet.append([term, value])

    # Saving the workbood

    workbook.save(workbook_path or "Translations.xlsx")

    print("done")


rogue_terms = {
    # Basics
    "ArcaneTrickster" : "Rogue",
    "CombatManeuver" : "Fighter",
    "Blowgun" : "Equipment",
    "Caltrops" : "Equipment",
    "Crossbow" : "Equipment",
    "CunningAction" : "Rogue",
    "Dagger" : "Equipment",
    "DragonResilience" : "Sorcerer",
    "EldritchKnight" : "Fighter",
    "ExtraAttack" : "Fighter",
    "Greataxe" : "Equipment",
    "HolySymbol" : "Equipment",
    "Invocation" : "Warlock",
    "Ladder" : "Equipment",
    "Lantern" : "Equipment",
    "Longbow" : "Equipment",
    "PactMagic" : "Warlock",
    "PactWeapon" : "Warlock",
    "PlateArmor" : "Equipment",
    "PotionOfHealing" : "Equipment",
    "Quarterstaff" : "Equipment",
    "SearUndead" : "Cleric",
    "SneakAttack" : "Rogue",
    "ThievesTools": "Equipment",
    "ThirstingBlade" : "Warlock",
    "Torch" : "Equipment",
    "TwoExtraAttacks" : "Fighter",
    "UnarmoredDefense" : "Barbarian",
    "Warhammer" : "Equipment",

    # Origins
    "CharismaModifier" : "Basics",
    "CharismaSavingThrow" : "Basics",
    "Cold" : "Basics",
    "CreatureType" : "Basics",
    "Duration" : "Basics",
    "Lightning" : "Basics",
    "Line:%:%" : "Basics",
    "NonHumanoid" : "Basics",
    "Prepared" : "Basics",
    "Range:%" : "Basics",
    "SkillProficiencies" : "Basics",
    "UnoccupiedSpace" : "Basics",
    "LowerPlanes": "Basics",
    "UpperPlanes": "Basics",
    "PlaneOfExistence": "Basics",

    # Equipment
    "Arcane" : "Basics",
    "AcidDamage" : "Basics",
    "CastingTime" : "Basics",
    "Cone:%" : "Basics",
    "DamageType" : "Basics",
    "Dehydration" : "Basics",
    "DexterityModifier" : "Basics",
    "DexterityScore" : "Basics",
    "Divine" : "Basics",
    "Expertise" : "Basics",
    "Fiend" : "Basics",
    "HP" : "Basics",
    "Humanoid" : "Basics",
    "Material" : "Basics",
    "MaterialComponent" : "Basics",
    "MaterialComponents" : "Basics",
    "OpportunityAttack" : "Basics",
    "Piercing": "Basics",
    "PoisonDamage" : "Basics",
    "Poison": "Basics",
    "Primal" : "Basics",
    "Psychic" : "Basics",
    "SpellAttackBonus" : "Basics",
    "Stabilize" : "Basics",
    "StrengthModifier" : "Basics",
    "ToolProficiency" : "Basics",
    "Undead" : "Basics",
    "Unwilling" : "Basics",
    "WeaponProficiencies" : "Basics",

    # Barbarian
    "Concentration": "Basics",
    "Emanation:%" : "Basics",
    "Enemy" : "Basics",
    "Force" : "Basics",
    "HitPointDie" : "Basics",
    "Hover" : "Basics",
    "Necrotic" : "Basics",
    "OuterPlanes" : "Basics",
    "Radiant" : "Basics",
    "Ritual" : "Basics",
    "SavingThrowProficiencies" : "Basics",
    "StrengthSavingThrow" : "Basics",
    "SubclassFeature" : "Basics", 

    # Bard
    "ToolProficiencies": "Basics",
}

terms = {}


workbook = openpyxl.Workbook()
if workbook_path and os.path.exists(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)

for sheet in workbook.worksheets:
    for row in sheet.rows:
        terms[row[0].value] = (row[1].value if len(row) > 1 else "", sheet.title)

for key, value in rogue_terms.items():
    terms[key] = (terms[key][0], value)

# Sorting the terms

terms = {term: value for term, value in sorted(terms.items(), key=lambda item: item[0])}



for sheet in workbook.worksheets:
    workbook.remove(sheet)

# Adding the terms to the workbook

for term, (value, sheet_name) in terms.items():
    if not sheet_name: sheet_name = "default"
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    sheet.append([term, value])

# Saving the workbood

workbook.save(workbook_path or "Translations.xlsx")