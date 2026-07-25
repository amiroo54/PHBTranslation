import re, os, sys
import openpyxl

if len(sys.argv) == 1:
    print("Usage: python term_extractor.py [DIRECTORY/FILE TO SCAN] [INPUT WORKBOOK] [SHEET NAME]")
    exit()

scan_path = sys.argv[1]
workbook_path = sys.argv[2] if len(sys.argv) > 2 else ""
sheet_name = sys.argv[3] if len(sys.argv) > 3 else ""

terms = []
directories = [scan_path]

# Extracting the terms from the directory/file

while directories:
    curr = directories.pop()
    if os.path.isdir(curr):
        for subdir in os.listdir(curr):
            directories.append(os.path.join(curr, subdir))
    if os.path.isfile(curr):
        if not curr.endswith(".md"):
            continue
        with open(curr, "r", encoding="utf-8") as file:
            matches = re.findall(r"\{\{(.*?)\}\}", file.read())
        for match in matches:
            parts = match.split(":")
            res = parts[0] + ":%" * (len(parts) - 1)
            terms.append(res)

terms = dict.fromkeys(sorted(terms), ("", sheet_name))

# Extracting the terms from the workbook

workbook = openpyxl.Workbook()
if workbook_path and os.path.exists(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)

for sheet in workbook.worksheets:
    for row in sheet.rows:
        terms[row[0].value] = (row[1].value if len(row) > 1 else "", sheet.title)

# Sorting the terms

terms = {term: value for term, value in sorted(terms.items(), key=lambda item: item[0])}

# Clearaing the sheets

for sheet in workbook.worksheets:
    workbook.remove(sheet)

# Adding the terms to the workbook

for term, (value, sheet_name) in terms.items():
    if not sheet_name: sheet_name = "default"
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    
    sheet = workbook[sheet_name]

    sheet.append([term, value])

# Saving the workbood

workbook.save(workbook_path or "Translations.xlsx")

print("done")